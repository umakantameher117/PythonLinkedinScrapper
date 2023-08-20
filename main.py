import openpyxl
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import InvalidArgumentException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Create a Chrome WebDriver instance
browser_options = webdriver.ChromeOptions()
driver = webdriver.Chrome(options=browser_options)

# Open the LinkedIn login page
driver.get('https://www.linkedin.com/login?fromSignIn=true&trk=guest_homepage-basic_nav-header-signin')

try:
    # Wait for the username field to be present and visible
    username = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//input[@id='username']"))
    )
    username.send_keys('youremail@email.com')

    # Find and fill the password field
    password = driver.find_element(By.XPATH, "//input[@id='password']")
    password.send_keys('yourPassword')

    # Find and click the login button
    login_button = driver.find_element(By.CLASS_NAME, "login__form_action_container")
    login_button.click()

    # Wait for manual verification before proceeding
    input("Please verify the sign-in process. Press Enter to continue...")

    # Initialize data dictionary
    data_dict = {}

    # Check if the file already exists
    excel_file = "linkedin_profiles.xlsx"
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Profile Name", "Profile Image URL", "Profile URL"])

    # Loop to input LinkedIn URLs
    while True:
        url = input("Enter a LinkedIn profile URL (or 'exit' to finish): ")

        if url.lower() == 'exit':
            break

        # Open the provided LinkedIn URL
        try:
            driver.get(url)

            # Wait for the profile name and profile image elements to be visible
            wait = WebDriverWait(driver, 10)
            profile_name_element = wait.until(
                EC.presence_of_element_located((By.XPATH, "//h1[@class='text-heading-xlarge inline t-24 v-align-middle break-words']"))
            )
            profile_image_element = wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, 'pv-top-card-profile-picture__image'))
            )

            # Extract the profile name, profile image URL, and profile URL
            profile_name = profile_name_element.text
            profile_image_url = profile_image_element.get_attribute("src")

            data_dict[url] = {
                "Profile Name": profile_name,
                "Profile Image URL": profile_image_url,
                "Profile URL": url
            }
        except InvalidArgumentException:
            print("Error: Invalid LinkedIn profile URL. Please enter a valid URL.")
            continue  # Continue to the next iteration
        except TimeoutException:
            print("Error: Timeout while loading the LinkedIn page. Please check the URL and try again.")
            continue  # Continue to the next iteration

except Exception as e:
    print("An error occurred:", e)

finally:
    # Close the browser
    driver.quit()

    # Append data to Excel sheet if any data was collected
    if data_dict:
        for url, profile_data in data_dict.items():
            sheet.append([profile_data["Profile Name"], profile_data["Profile Image URL"], profile_data["Profile URL"]])

        workbook.save(excel_file)
        print("Data appended to 'linkedin_profiles.xlsx'")
    else:
        print("No data was collected.")
